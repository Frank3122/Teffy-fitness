from django.shortcuts import render,redirect,get_object_or_404
from teffyapp.models import *
from django.contrib.auth import login,logout,authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from enum import Enum
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.http import HttpResponseForbidden
from django.utils import timezone
from datetime import date, datetime
from django.db.models import Sum, F
from django.utils.timezone import now  
from django.utils.timezone import localtime
from datetime import datetime , timedelta
from django.http import HttpResponseRedirect, HttpResponse, Http404
from decimal import Decimal
import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.templatetags.static import static
# from teffyapp.models import AddMember, PersonalInformation, Payments
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
import os
# from .models import PersonalInformation, Payments
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
# from django.shortcuts import render
from django.core.exceptions import BadRequest
from django.db.models import Q


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils.dateparse import parse_date
from .models import AddMember, Service, UserProfile, Plan

from django.db.models import Sum
from django.contrib import messages


from datetime import date
from dateutil.relativedelta import relativedelta
from django.shortcuts import render
from teffyapp.models import PersonalInformation


# Create your views here.

class Messages(Enum):
    SUCCESS = 1
    ERROR = 2
    WARNING = 3
    INFO = 4 

def success_message() -> str:
    return "Success message from success function"

def error_message() -> str:
    return "Error message from error function"

def warning_message() -> str:
    return "Warning message from warning function"

def info_message() -> str:
    return "Info message from info function"

def user_login(request):
    if request.method == "POST":
        try:
            username = request.POST.get("username")
            password = request.POST.get("password")
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)  
                # messages.success(request, "WELCOME TO THE PAGE")
                return redirect("/main/")  
            else:
                messages.error(request, "Invalid username or password. Please try again.")
                return redirect("/login/")  
            
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect("/login/")
        
    return render(request, "login.html")


def user_logout(request):
    logout(request)
    return redirect("/login/")

def register(request):
    if request.method == "POST":
        try:
            username = request.POST.get("username")
            password = request.POST.get("password")
            confirm_password = request.POST.get("confirm_password")
            email = request.POST.get("email")
            if User.objects.filter(username__iexact=username).exists():
                messages.error(request, "Username already exists.")
                return redirect("/register/")
            if password != confirm_password:
                messages.error(request, "Passwords do not match.")
                return redirect("/register/")
            user = User.objects.create_user(username=username, email=email)
            user.set_password(password)
            user.save()
            messages.success(request, "Successfully Registered! Please login.")
            return redirect("/login/")  
        
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect("/register/")
        
    return render(request, "register.html")

def forms(request):
    if request.method == "POST":
        try:
            name = request.POST.get("name")
            address = request.POST.get("address")
            occupation = request.POST.get("occupation")
            gender = request.POST.get("gender")
            mobile_number = request.POST.get("mobile_number")
            email = request.POST.get("email")
            perfect_body = request.POST.get("perfect_body")
            body_type = request.POST.get("body_type")
            goal = request.POST.get("goal")
            body_you_want = request.POST.get("body_you_want")
            level_of_body_fat = request.POST.get("level_of_body_fat")
            problem_areas = request.POST.get("problem_areas")
            diets = request.POST.get("diets")
            water_you_drink_daily = request.POST.get("water_you_drink_daily")
            eat_or_dividie_foods_or_beverages = request.POST.get("eat_or_dividie_foods_or_beverages")
            event_coming_up = request.POST.get("event_coming_up")
            height = request.POST.get("height")
            current_weight = request.POST.get("current_weight")
            target_weight = request.POST.get("target_weight")
            level_of_fitness = request.POST.get("level_of_fitness")
            status = request.POST.get("status", "new")

            print("Form Data:", {"name": name, "status": status})

            personal_info = PersonalInformation(
                name=name,
                address=address,
                occupation=occupation,
                gender=gender,
                mobile_number=mobile_number,
                email=email,
                perfect_body=perfect_body,
                body_type=body_type,
                goal=goal,
                body_you_want=body_you_want,
                level_of_body_fat=level_of_body_fat,
                problem_areas=problem_areas,
                diets=diets,
                water_you_drink_daily=water_you_drink_daily,
                eat_or_dividie_foods_or_beverages=eat_or_dividie_foods_or_beverages,
                event_coming_up=event_coming_up,
                height=height,
                current_weight=current_weight,
                target_weight=target_weight,
                level_of_fitness=level_of_fitness,
                status=status,
            )
            personal_info.save()

            return redirect(request.path)

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect(request.path)

    return render(request, "forms.html")



# def add_leads(request):
#     if request.method == "POST":
       
#         name = request.POST.get("name")
#         address = request.POST.get("address")
#         occupation = request.POST.get("occupation")
#         gender = request.POST.get("gender")
#         mobile_number = request.POST.get("mobile_number")
#         email = request.POST.get("email")
#         perfect_body = request.POST.get("perfect_body")
#         body_type = request.POST.get("body_type")
#         goal = request.POST.get("goal")
#         body_you_want = request.POST.get("body_you_want")
#         level_of_body_fat = request.POST.get("level_of_body_fat")
#         problem_areas = request.POST.get("problem_areas")
#         diets = request.POST.get("diets")
#         water_you_drink_daily = request.POST.get("water_you_drink_daily")
#         eat_or_dividie_foods_or_beverages = request.POST.get("eat_or_dividie_foods_or_beverages")
#         event_coming_up = request.POST.get("event_coming_up")
#         height = request.POST.get("height")
#         current_weight = request.POST.get("current_weight")
#         target_weight = request.POST.get("target_weight")
#         level_of_fitness = request.POST.get("level_of_fitness")
#         status = request.POST.get("status", 'new')  

#         print("Form Data:", { 
#             "name": name,
#             "status": status,
#         })

        
#         personal_info = PersonalInformation(
#             name=name,
#             address=address,
#             occupation=occupation,
#             gender=gender,
#             mobile_number=mobile_number,
#             email=email,
#             perfect_body=perfect_body,
#             body_type=body_type,
#             goal=goal,
#             body_you_want=body_you_want,
#             level_of_body_fat=level_of_body_fat,
#             problem_areas=problem_areas,
#             diets=diets,
#             water_you_drink_daily=water_you_drink_daily,
#             eat_or_dividie_foods_or_beverages=eat_or_dividie_foods_or_beverages,
#             event_coming_up=event_coming_up,
#             height=height,
#             current_weight=current_weight,
#             target_weight=target_weight,
#             level_of_fitness=level_of_fitness,
#             status=status,  
#         )
#         personal_info.save()

#         return redirect(request.path)

#     return render(request, "adding-leads.html")


def main(request):
    # Get today's date
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)


    pending_members = AddMember.objects.annotate(
        pending_amount=F('total_amount') - F('current_installment_amount')
    ).filter(pending_amount__gt=0)

    # Today's enrollments
    today_enrollments = PersonalInformation.objects.filter(
        status='converted', 
        assigned_date__year=today.year,
        assigned_date__month=today.month,
        assigned_date__day=today.day
    ).count()

    today_add_members_count = AddMember.objects.filter(
        enrollment_date__year=today.year,
        enrollment_date__month=today.month,
        enrollment_date__day=today.day
    ).count()

    daily_enrollments = today_enrollments + today_add_members_count

    # Lead counts
    member_count = PersonalInformation.objects.filter(status='converted').count()
    all_member_count = PersonalInformation.objects.count() + AddMember.objects.count()
    not_converted_count = PersonalInformation.objects.filter(status='not_converted').count()
    pending_leads_count = PersonalInformation.objects.filter(status="pending").count()
    new_leads_count = PersonalInformation.objects.filter(status="new").count()
    total_members = AddMember.objects.count()

    # Other counts
    staffs_count = UserProfile.objects.all().count()
    stocks_count = Stock.objects.all().count()
    returns_count = Returns.objects.all().count()

    # Sales count
    sales = Sales.objects.filter(sale_date__year=today.year, sale_date__month=today.month).count()
    today_sales = Sales.objects.filter(
        sale_date__year=today.year,
        sale_date__month=today.month,
        sale_date__day=today.day
    ).count()

    # Purchase count
    purchase = Purchase.objects.filter(purchase_date__year=today.year, purchase_date__month=today.month).count()
    today_purchase = Purchase.objects.filter(
        purchase_date__year=today.year,
        purchase_date__month=today.month,
        purchase_date__day=today.day
    ).count()

    # Follow-up leads
    follow_up = PersonalInformation.objects.filter(status='follow_up').count()
    today_follow_up = PersonalInformation.objects.filter(
        status="follow_up", 
        assigned_date__year=today.year,
        assigned_date__month=today.month,
        assigned_date__day=today.day
    ).count()

    # Monthly enrollments
    converted_count = PersonalInformation.objects.filter(
        status='converted', 
        assigned_date__year=today.year, 
        assigned_date__month=today.month
    ).count()

    enrolled_count = AddMember.objects.filter(
        enrollment_date__year=today.year, 
        enrollment_date__month=today.month
    ).count()
    
    total_enrollments = converted_count + enrolled_count

    # Renewals
    today_renewals = Renew.objects.filter(
        renew_date__year=today.year,
        renew_date__month=today.month,
        renew_date__day=today.day
    ).count()

    renewals = Renew.objects.filter(
        renew_date__year=today.year,
        renew_date__month=today.month
    ).count()

    # Expiry Dates
    expiry_date = Renew.objects.filter(
        expiry_date__year=today.year,
        expiry_date__month=today.month
    ).count()

    today_expiry_date = Renew.objects.filter(
        expiry_date__year=today.year,
        expiry_date__month=today.month,
        expiry_date__day=today.day
    ).count()

    # Revenue Calculation (Sum of Payment Amounts)
    total_revenue = Payments.objects.aggregate(total=Sum('amount_paid'))['total'] or 0
    today_revenue = Payments.objects.filter(date_paid__date=today).aggregate(total=Sum('amount_paid'))['total'] or 0

    # Total expense calculation
    expenses = Expense.objects.all().annotate(total_price=F('price') * F('quantity'))

    total_expense = Expense.objects.filter(date_spent__gte=first_day_of_month, date_spent__lt=today.replace(day=1) + timedelta(days=32)).aggregate(total=Sum('price'))['total'] or 0
    today_expense = Expense.objects.filter(date_spent=today).aggregate(total=Sum(F('price') * F('quantity')))['total'] or 0

    # Total profit calculation
    total_profit = total_revenue - total_expense
    today_profit = today_revenue - today_expense

    # Overdue follow-ups
    overdue_follow_ups = PersonalInformation.objects.filter(
        follow_up_date__lte=today, status="follow_up"
    )

    # **Pass everything as context**
    context = {
        "member_count": member_count,
        "sales": sales,
        "purchase": purchase,
        "follow_up": follow_up,
        "all_member_count": all_member_count,
        "total_enrollments": total_enrollments,
        "renewals": renewals,
        "daily_enrollments": daily_enrollments,
        "today_renewals": today_renewals,
        "today_sales": today_sales,
        "today_purchase": today_purchase,
        "today_follow_up": today_follow_up,
        "expiry_date": expiry_date,
        "today_expiry_date": today_expiry_date,
        # "today_revenue": today_revenue,
        "not_converted_count": not_converted_count,
        "pending_leads_count": pending_leads_count,
        "new_leads_count": new_leads_count,
        "staffs_count": staffs_count,
        "stocks_count": stocks_count,
        "returns_count": returns_count,
        "overdue_follow_up_count": overdue_follow_ups.count(),
        "overdue_follow_ups": overdue_follow_ups,
        # 'today_profit':today_profit,
        # 'today_expense':today_expense,
        'total_members':total_members,
        "pending_members": pending_members,  # Pass to the dashboard
        "pending_count": pending_members.count(),

        "total_revenue": total_revenue,
        "today_revenue": today_revenue,
        "total_expense": total_expense,
        "today_expense": today_expense,
        "total_profit": total_profit,
        "today_profit": today_profit,
        "expenses":expenses,
    }

    # Success message
    messages.success(request, "WELCOME TO TEFFY FITNESS")

    # Render the response
    return render(request, 'main-page.html', context)


def expense(request):
    return render(request,"expense-page.html")

def follow_up(request):
    return render(request,"follow-up page.html")

def plan(request):
    if request.method == "POST":
        try:
            plan_name = request.POST.get('plan_name')
            price = request.POST.get('price')
            duration = request.POST.get('duration')
            batch = request.POST.get('batch')

            if plan_name and price and duration:
                Plan.objects.create(plan_name=plan_name, price=price, duration=duration, batch=batch)
                return redirect('plan')  

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('plan')

    all_plans = Plan.objects.all()
    return render(request, "plan-page.html", {'all_plans': all_plans})


def edit_plan(request, plan_id):
    try:
        obj = get_object_or_404(Plan, id=plan_id)

        if request.method == "POST":
            plan_name = request.POST.get('plan_name')
            price = request.POST.get('price')
            duration = request.POST.get('duration')
            batch = request.POST.get('batch')

            if plan_name:
                obj.plan_name = plan_name
            if price:
                obj.price = price
            if duration:
                obj.duration = duration
            if batch:
                obj.batch = batch

            obj.save()
            messages.success(request, "Plan updated successfully!")
            return redirect(f'/plan/{obj.id}/')  

        data = obj.plan()
        return render(request, "edit-plan.html", data)

    except Plan.DoesNotExist:
        messages.error(request, "Plan not found!")
        return redirect('/plans/')  # Redirect to a plans list or appropriate page

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('/plans/')  # Redirect to avoid displaying raw errors
    


def assign_plan(request, id):
    if request.method == "POST":
        try:
            plan_id = request.POST.get("plan")  # Ensure correct field name
            print(f"Received Plan ID: {plan_id}")  # Debugging
            
            plan = Plan.objects.filter(id=plan_id).first()
            lead = PersonalInformation.objects.filter(id=id).first()
            print(f"Lead: {lead}, Plan: {plan}")  # Debugging

            if plan and lead:
                print(f"Assigning Plan {plan.plan_name} to Lead {lead.id}")  # Debugging
                lead.plan_leads = plan
                lead.save()
                messages.success(request, "Plan successfully assigned!")
            else:
                messages.error(request, "Invalid Plan or Lead.")

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
        
        return redirect('converted_leads')  # Ensure the page refreshes
    
    
def delete_plan(request, plan_id):
    try:
        plan = get_object_or_404(Plan, id=plan_id)
        plan.delete()
        messages.success(request, "Plan successfully deleted!")
    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
    
    return redirect('plan')



# def price(request):
#     return render(request,"price-page.html")

def workplan(request):
    return render(request,"workplan.html")

def personal_info_list(request):
    all_data = PersonalInformation.objects.all()
    return render(request, 'new-leads.html', {'all_data': all_data})
    
def getone_lead(request, id):
    try:
        obj = PersonalInformation.objects.get(id=id)
        data = obj.gym()
    except PersonalInformation.DoesNotExist:
        messages.error(request, "The requested lead does not exist.")
        return redirect('error_page')  
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('error_page')

    print("ID is available:", obj.id)  # Debugging

    if request.method == "POST":
        try:
            name = request.POST.get("name")
            address = request.POST.get("address")
            occupation = request.POST.get("occupation")
            gender = request.POST.get("gender")
            mobile_number = request.POST.get("mobile_number")
            email = request.POST.get("email")
            perfect_body = request.POST.get("perfect_body")
            body_type = request.POST.get("body_type")
            goal = request.POST.get("goal")
            body_you_want = request.POST.get("body_you_want")
            level_of_body_fat = request.POST.get("level_of_body_fat")
            problem_areas = request.POST.get("problem_areas")
            diets = request.POST.get("diets")
            water_you_drink_daily = request.POST.get("water_you_drink_daily")
            eat_or_dividie_foods_or_beverages = request.POST.get("eat_or_dividie_foods_or_beverages")
            event_coming_up = request.POST.get("event_coming_up")
            height = request.POST.get("height")
            current_weight = request.POST.get("current_weight")
            target_weight = request.POST.get("target_weight")
            level_of_fitness = request.POST.get("level_of_fitness")
            status = request.POST.get("status")

            # Update fields if provided
            if name:
                obj.name = name
            if address:
                obj.address = address
            if occupation:
                obj.occupation = occupation
            if gender:
                obj.gender = gender
            if mobile_number:
                obj.mobile_number = mobile_number
            if email:
                obj.email = email
            if perfect_body:
                obj.perfect_body = perfect_body
            if body_type:
                obj.body_type = body_type
            if goal:
                obj.goal = goal
            if body_you_want:
                obj.body_you_want = body_you_want
            if level_of_body_fat:
                obj.level_of_body_fat = level_of_body_fat
            if problem_areas:
                obj.problem_areas = problem_areas
            if diets:
                obj.diets = diets
            if water_you_drink_daily:
                obj.water_you_drink_daily = water_you_drink_daily
            if eat_or_dividie_foods_or_beverages:
                obj.eat_or_dividie_foods_or_beverages = eat_or_dividie_foods_or_beverages
            if event_coming_up:
                obj.event_coming_up = event_coming_up
            if height:
                obj.height = height
            if current_weight:
                obj.current_weight = current_weight
            if target_weight:
                obj.target_weight = target_weight
            if level_of_fitness:
                obj.level_of_fitness = level_of_fitness

            # Always update status
            if status:
                obj.status = status

            obj.save()
            messages.success(request, "Data updated successfully")
            return redirect(f'/leads/{obj.id}/')

        except Exception as e:
            messages.error(request, f"An error occurred while updating data: {e}")
            return redirect(f'/leads/{obj.id}/')

    return render(request, "one-leads.html", data)


def update_followup_date(request, id):
    """Update follow-up date when changed in the dropdown"""
    if request.method == "POST":
        try:
            lead = get_object_or_404(PersonalInformation, id=id)
            follow_up_date = request.POST.get("follow_up_date")

            if follow_up_date:
                lead.follow_up_date = follow_up_date
                lead.save()
                messages.success(request, f"Follow-up date updated for {lead.name}")
            else:
                messages.warning(request, "No follow-up date provided.")

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

    return redirect("follow_up_leads")  # Redirect back to leads page


def update_status(request, id):
    if request.method == 'POST':
        try:
            lead = PersonalInformation.objects.get(id=id)
            status = request.POST.get('status')

            if status:
                lead.status = status
                lead.save()
                messages.success(request, f"Status updated to {status} successfully.")

                if status == 'new':
                    return redirect('new_leads')
                elif status == 'converted':
                    return redirect('converted_leads')
                elif status == 'not_converted':
                    return redirect('not_converted_leads')
                elif status == 'pending':
                    return redirect('pending_leads')
                elif status == 'follow_up':
                    return redirect('follow_up_leads')
            else:
                messages.warning(request, "No status provided.")
        
        except PersonalInformation.DoesNotExist:
            messages.error(request, "Lead not found.")
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

    return redirect('new_leads')


def new_leads(request):
    all_data = PersonalInformation.objects.filter(status='new') 
    # add_members = AddMember.objects.all()
    return render(request, 'new-leads.html', {'all_data': all_data, 'status_title': 'New',})


def converted_leads(request):
    if request.method == "POST":
        try:
            lead_id = request.POST.get('lead_id')
            service_id = request.POST.get('services')
            plan_id = request.POST.get('plan_leads')

            lead = PersonalInformation.objects.get(id=lead_id)
            selected_service = Service.objects.get(id=service_id)
            selected_plan = Plan.objects.get(id=plan_id)

            # Assign service and plan
            lead.services = selected_service
            lead.plan_leads = selected_plan
            lead.status = "converted"  # Ensure status is updated
            lead.save()

            # Ensure no duplicate AddMember creation
            if not AddMember.objects.filter(email=lead.email).exists():
                last_member = AddMember.objects.order_by('-id').first()
                last_id = int(last_member.member_id.replace("TFM", "")) if last_member and last_member.member_id else 0
                new_member_id = f"TFM{last_id + 1}"

                AddMember.objects.create(
                    member_id=new_member_id,
                    name=lead.name,
                    mobile_number=lead.mobile_number,
                    email=lead.email,
                    converted_from_lead=lead,  # Link the lead to this member
                    service=selected_service,
                    select_membership_plan=selected_plan
                )

            messages.success(request, f"Lead {lead.name} successfully converted.")
        
        except PersonalInformation.DoesNotExist:
            messages.error(request, "Lead not found.")
        except Service.DoesNotExist:
            messages.error(request, "Selected service not found.")
        except Plan.DoesNotExist:
            messages.error(request, "Selected plan not found.")
        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {e}")

        return redirect('converted_leads')

    try:
        all_data = PersonalInformation.objects.filter(status='converted')
        available_services = Service.objects.all()
        plans = Plan.objects.all()

    except Exception as e:
        messages.error(request, f"Error loading data: {e}")
        return redirect('error_page')

    return render(request, 'converted_leads.html', {
        'all_data': all_data,
        'status_title': 'Converted',
        'available_services': available_services,
        'plans': plans,
    })
    
    
def update_assigned_date(request, id):
    if request.method == "POST":
        assigned_date = request.POST.get("assigned_date")
        lead = get_object_or_404(PersonalInformation, id=id)

        if assigned_date:
            lead.assigned_date = assigned_date
            lead.save()
            messages.success(request, f"Assigned date updated for {lead.name}!")

    return redirect("converted_leads")



def not_converted_leads(request):
    if request.method == "POST":
        try:
            lead_id = request.POST.get("lead_id")  # Get the lead ID from the form
            lead = get_object_or_404(PersonalInformation, id=lead_id)

            # Update the status
            status = request.POST.get("status")
            lead.status = status

            # If the status is 'pending', store the reason; otherwise, clear it
            if status == "pending":
                lead.reason = request.POST.get("reason", "")
            else:
                lead.reason = ""

            lead.save()
            messages.success(request, f"Lead {lead.name} status updated successfully.")
            return redirect("pending_leads")  # Refresh the page after updating

        except PersonalInformation.DoesNotExist:
            messages.error(request, "Lead not found.")
        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {e}")

    try:
        all_data = PersonalInformation.objects.filter(status='not_converted')
    except Exception as e:
        messages.error(request, f"Error loading data: {e}")
        return redirect("error_page")  # Redirect to a generic error page if needed

    return render(request, 'not_converted_leads.html', {
        'all_data': all_data,
        'status_title': 'Not Converted'
    })


def pending_leads(request):
    if request.method == "POST":
        try:
            lead_id = request.POST.get("lead_id")  # Get the lead ID from the form
            lead = get_object_or_404(PersonalInformation, id=lead_id)

            # Update the status
            status = request.POST.get("status")
            lead.status = status

            # If the status is 'pending', store the reason; otherwise, clear it
            if status == "pending":
                lead.reason = request.POST.get("reason", "")
            else:
                lead.reason = ""

            lead.save()
            messages.success(request, f"Lead {lead.name} status updated successfully.")
            return redirect("pending_leads")  # Refresh the page after updating

        except PersonalInformation.DoesNotExist:
            messages.error(request, "Lead not found.")
        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {e}")

    try:
        # Get all leads with 'pending' status
        all_data = PersonalInformation.objects.filter(status="pending")

    except Exception as e:
        messages.error(request, f"Error loading data: {e}")
        return redirect("pending_leads")  # Reloads the same page with an error message

    return render(request, "pending_leads.html", {
        "all_data": all_data,
        "status_title": "Pending"
    })


def follow_up_leads(request):
    all_data = PersonalInformation.objects.filter(status='follow_up')
    return render(request, 'follow_up_leads.html', {'all_data': all_data, 'status_title': 'Follow Up'})

def delete_lead(request, id):
    lead = PersonalInformation.objects.get(id=id)
    lead.delete()
    messages.success(request, "Lead deleted successfully.")
    return redirect('converted_leads') ############


def user_management(request):
    try:
        if request.method == "POST":
            profile_user_id = request.POST.get("profile_user_id")
            username = request.POST.get("name")
            password = request.POST.get("password")
            group = request.POST.get("group")

            if not username or not password or not group:
                messages.error(request, "All fields are required.")
                return redirect("/user_management/")

            if User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.")
                return redirect("/user_management/")

            if UserProfile.objects.filter(profile_user_id=profile_user_id).exists():
                messages.error(request, "Profile User ID already exists.")
                return redirect("/user_management/")

            # Create user and set password
            user = User.objects.create_user(username=username)
            user.set_password(password)

            # Assign roles
            if group == "admin":
                user.is_superuser = True
            elif group == "staff":
                user.is_staff = True
                user.is_superuser = False

            user.save()

            # Auto-generate profile_user_id if not provided
            if not profile_user_id:
                user_count = UserProfile.objects.count() + 1
                profile_user_id = f"TFS-{user_count}"

            # Create user profile
            UserProfile.objects.create(
                user=user,
                profile_user_id=profile_user_id,
                group=group,
            )

            messages.success(request, "User successfully added.")
            return redirect("/user_management/")

        # Fetch user data
        user_count = UserProfile.objects.count() + 1
        profile_user_id = f"TFS-{user_count}"
        data = UserProfile.objects.all()

        # Check user roles
        is_admin = request.user.is_superuser
        is_staff = request.user.is_staff

        context = {
            "data": data,
            "profile_user_id": profile_user_id,
            "is_admin": is_admin,
            "is_staff": is_staff,
        }

        return render(request, "user-management.html", context)

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect("/user_management/")



def edit_user(request, user_id):
    try:
        user_profile = UserProfile.objects.get(profile_user_id=user_id)

        if request.method == 'POST':
            username = request.POST.get('name')
            group = request.POST.get('group')
            password = request.POST.get('password')

            if not username or not group:
                messages.error(request, "Username and Group fields cannot be empty.")
                return redirect('edit_user', user_id=user_id)

            user_profile.user.username = username
            user_profile.group = group

            if password:
                user_profile.user.set_password(password)

            user_profile.user.save()
            user_profile.save()

            messages.success(request, "User details successfully updated.")
            return redirect('user_management')

        return render(request, 'edit-user.html', {'user_profile': user_profile})

    except UserProfile.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect('user_management')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('user_management')



def delete_user(request, user_id):
    user_profile = UserProfile.objects.get(profile_user_id = user_id)
    user_profile.user.delete()  
    user_profile.delete()       
    
    messages.success(request, "User successfully deleted.")
    return redirect('user_management')

def view_all_staffs(request):
    admins = UserProfile.objects.filter(group="admin")
    staff = UserProfile.objects.filter(group="staff")

    context = {
        "admins": admins,
        "staff": staff,
    }
    return render(request, "view-all-staffs.html", context)

def services(request):
    try:
        if request.method == "POST":
            service_name = request.POST.get('service_name')
            prices = request.POST.get("price")
            duration = request.POST.get("duration")
            group = request.POST.get("group")
            sessions = request.POST.get("sessions")

            if not service_name or not prices or not duration or not group or not sessions:
                messages.error(request, "All fields are required.")
                return redirect('services')

            Service.objects.create(
                name=service_name,
                prices=prices,
                duration=duration,
                group=group,
                sessions=sessions
            )
            messages.success(request, "Service successfully added.")
            return redirect('services')

        all_services = Service.objects.all()
        return render(request, 'manage-services.html', {'services': all_services})

    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
        return redirect('services')


def delete_service(request, service_id):
    service = Service.objects.get(id=service_id)
    service.delete()
    return redirect('services')



def assign_service(request, id):
    try:
        if request.method == "POST":
            service_id = request.POST.get("service")

            service = Service.objects.filter(id=service_id).first()
            lead = PersonalInformation.objects.filter(id=id).first()

            if service and lead:
                lead.services = service
                lead.save()
                messages.success(request, "Service successfully assigned!")
            else:
                messages.error(request, "Invalid service or lead.")

        return redirect('converted_leads')

    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
        return redirect('converted_leads')



def overall_services(request):
        client_details = PersonalInformation.objects.filter(status='converted')
        print("Number of clients:", client_details.count())
        for client in client_details:
            print(f"Client: {client.name}, Service: {client.services.name if client.services else 'No Service'}")
        services = Service.objects.all()

        context = {
            "client_details": client_details, 
            "services_taken": services,
        }
        return render(request, "overall-service.html", context)


def payment(request, client_id):
    try:
        client = get_object_or_404(PersonalInformation, id=client_id)

        # Get the service price if it exists
        service_price = client.services.prices if client.services else 0  

        # Get the membership price if a membership plan exists
        membership_price = client.plan_leads.price if client.plan_leads else 0  

        # Calculate the total price (service price + membership price)
        total_price = service_price + membership_price

        # Get the total amount already paid by the client
        total_amount_paid = Payments.objects.filter(payments=client).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0

        # Calculate pending amount
        pending_amount = total_price - total_amount_paid

        # ✅ Fetch expiry date from PersonalInformation
        expiry_date = client.expiry_date if client.expiry_date else "Not Available"

        # ✅ Debugging: Print expiry date in terminal
        print(f"DEBUG: Expiry Date for {client.name} ({client.id}): {expiry_date}")

        if request.method == 'POST':
            try:
                amount_paid = float(request.POST.get('amount_paid', 0))
                payment_mode = request.POST.get('payment_mode')

                if amount_paid > 0 and payment_mode:
                    new_pending_amount = pending_amount - amount_paid  

                    # Save payment details
                    Payments.objects.create(
                        payments=client,
                        name=client,
                        amount_paid=amount_paid,
                        pending_amount=new_pending_amount,
                        payment_mode=payment_mode,
                        date_paid=now(),  
                    )
                    messages.success(request, "Payment submitted successfully!")
                    return redirect('payment_page', client_id=client_id)

                messages.error(request, "Invalid payment details. Please check the amount and payment mode.")

            except Exception as e:
                messages.error(request, f"An error occurred while processing the payment: {e}")
                return redirect('payment_page', client_id=client_id)

        # Get payment history
        payment_history = Payments.objects.filter(payments=client)

        return render(request, 'payment.html', {
            'client': client,
            'service_price': service_price,
            'membership_price': membership_price,  # Pass membership price to template
            'total_price': total_price,  # Total price (Service + Membership)
            'total_amount_paid': total_amount_paid,
            'pending_amount': pending_amount,
            'payment_history': payment_history,
            'expiry_date': expiry_date,  # ✅ Ensure expiry_date is passed
        })

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('payment_page', client_id=client_id)

def expenses(request):
    current_year = datetime.now().year
    current_month = datetime.now().month

    years = list(range(current_year - 5, current_year + 1))

    months = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August",
        9: "September", 10: "October", 11: "November", 12: "December"
    }

    selected_year = int(request.GET.get("year", current_year))
    selected_month = int(request.GET.get("month", current_month))

    expenses = Expense.objects.filter(date_spent__year=selected_year, date_spent__month=selected_month)
    total_monthly_expenses = expenses.aggregate(Sum("price"))["price__sum"] or 0

    return render(request, "new-expense.html", {
        "expenses": expenses,
        "total_monthly_expenses": total_monthly_expenses,
        "years": years,
        "months": months,
        "selected_year": selected_year,
        "selected_month": selected_month,
    })


def add_expense(request):
    try:
        if request.method == 'POST':
            expense_name = request.POST.get('expense_name')
            price = request.POST.get('price')
            date_spent = request.POST.get('date_spent')
            quantity = request.POST.get('quantity')

            # Ensure required fields are provided
            if not expense_name or not price or not date_spent or not quantity:
                messages.error(request, "All fields are required.")
                return redirect('expenses')

            try:
                # Convert price and quantity to proper types
                price = float(price)
                quantity = int(quantity)

                # Create and save expense
                expense = Expense(expense_name=expense_name, price=price, date_spent=date_spent, quantity=quantity)
                expense.save()

                messages.success(request, "Expense added successfully.")
                return redirect('expenses')

            except ValueError:
                messages.error(request, "Invalid input for price or quantity.")
                return redirect('expenses')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('expenses')

    return render(request, 'your_template.html')  # Replace with your actual template name


def delete_expense(request, expense_id):
    expense = Expense.objects.get(id=expense_id)
    expense.delete()
    return redirect('expenses')

def view_expense(request, expense_id):
    try:
        obj = Expense.objects.get(id=expense_id)
        data = obj.expense()

        if request.method == "POST":
            try:
                expense_name = request.POST.get('expense_name')
                price = request.POST.get('price')
                date_spent = request.POST.get('date_spent')

                if expense_name:
                    obj.expense_name = expense_name

                if price:
                    obj.price = float(price)  # Ensure price is a valid number

                if date_spent:
                    obj.date_spent = date_spent

                obj.save()
                messages.success(request, "Expense updated successfully.")
                return redirect(f'/expenses/{obj.id}/')

            except ValueError:
                messages.error(request, "Invalid price format.")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

    except Expense.DoesNotExist:
        messages.error(request, "Expense not found.")
        return redirect('expenses')  # Redirect to the expense list page

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('expenses')

    return render(request, "view-expense.html", data)

        

def purchase(request):
    try:
        if request.method == "POST":
            try:
                # Get data from form
                product_name = request.POST.get("product_name")
                brand = request.POST.get("brand")
                quantity = request.POST.get("quantity")
                amount = request.POST.get("amount")
                purchase_date = request.POST.get("purchase_date")

                # Ensure all required fields are present
                if not all([product_name, brand, quantity, amount, purchase_date]):
                    messages.error(request, "All fields are required.")
                    return redirect("purchase")

                # Convert numeric fields
                quantity = int(quantity)
                amount = float(amount)

                # Create purchase record
                Purchase.objects.create(
                    product_name=product_name,
                    brand=brand,
                    quantity=quantity,
                    amount=amount,
                    purchase_date=purchase_date
                )

                messages.success(request, "Purchase record added successfully.")
                return redirect("purchase")

            except ValueError:
                messages.error(request, "Invalid quantity or amount format.")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

        # Retrieve all purchases
        purchase = Purchase.objects.all()

        # Debugging: Check if purchase_date is retrieved correctly
        for item in purchase:
            print(item.purchase_date)

        return render(request, "purchase.html", {"purchase": purchase})

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect("purchase")


def edit_purchase(request, purchase_id):
    try:
        obj = Purchase.objects.get(id=purchase_id)

        if request.method == "POST":
            try:
                product_name = request.POST.get("product_name")
                brand = request.POST.get("brand")
                quantity = request.POST.get("quantity")
                amount = request.POST.get("amount")
                purchase_date = request.POST.get("purchase_date")

                # Validate and update fields
                if product_name:
                    obj.product_name = product_name
                if brand:
                    obj.brand = brand
                if quantity:
                    obj.quantity = int(quantity)  # Ensure integer format
                if amount:
                    obj.amount = float(amount)  # Ensure float format
                if purchase_date:
                    obj.purchase_date = purchase_date

                obj.save()
                messages.success(request, "Purchase details updated successfully.")
                return redirect('purchase')

            except ValueError:
                messages.error(request, "Invalid quantity or amount format.")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

        data = {"purchase": obj}
        return render(request, "edit-purchase.html", data)

    except Purchase.DoesNotExist:
        messages.error(request, "Purchase record not found.")
        return redirect('purchase')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('purchase')


def delete_purchase(request, purchase_id):
    purchase = Purchase.objects.get(id=purchase_id)
    purchase.delete()
    return redirect('purchase')

def sales(request):
    try:
        if request.method == 'POST':
            product_name = request.POST.get('product_name')
            customer = request.POST.get('customer')
            quantity = request.POST.get('quantity')
            sale_price = request.POST.get('sale_price')
            sale_date = request.POST.get('sale_date')

            if not all([product_name, customer, quantity, sale_price, sale_date]):
                messages.error(request, "All fields are required.")
                return redirect('sales')

            try:
                product = Purchase.objects.get(product_name=product_name)

                # Ensure numeric values are valid
                quantity = int(quantity)
                sale_price = float(sale_price)

                Sales.objects.create(
                    product=product,
                    customer=customer,
                    quantity=quantity,
                    sale_price=sale_price,
                    sale_date=sale_date,
                )

                messages.success(request, "Sale recorded successfully.")
                return redirect('sales')

            except Purchase.DoesNotExist:
                messages.error(request, "Product not found in inventory.")
            except ValueError:
                messages.error(request, "Invalid quantity or sale price format.")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {e}")

        purchased_products = Purchase.objects.values('product_name').distinct()
        sales = Sales.objects.all()
        return render(request, 'sales.html', {'sales': sales, 'purchased_products': purchased_products})

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('sales')


def edit_sales(request, sales_id):
    try:
        obj = Sales.objects.get(id=sales_id)
        
        if request.method == "POST":
            product_name = request.POST.get('product_name')
            customer = request.POST.get('customer')
            quantity = request.POST.get('quantity')
            sale_price = request.POST.get('sale_price')
            sale_date = request.POST.get('sale_date')

            try:
                if product_name:
                    purchase_obj = Purchase.objects.get(product_name=product_name)
                    obj.product = purchase_obj 

                if customer:
                    obj.customer = customer

                if quantity:
                    obj.quantity = int(quantity)  # Ensure valid integer input

                if sale_price:
                    obj.sale_price = float(sale_price)  # Ensure valid float input

                if sale_date:
                    obj.sale_date = sale_date

                obj.save()
                messages.success(request, "Sale record updated successfully.")
                return redirect('sales')

            except Purchase.DoesNotExist:
                messages.error(request, "Product not found in inventory.")
            except ValueError:
                messages.error(request, "Invalid quantity or sale price format.")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {e}")

        purchased_products = Purchase.objects.all()
        data = {
            "sales": obj,
            "purchased_products": purchased_products,
        }
        return render(request, "edit-sales.html", data)

    except Sales.DoesNotExist:
        messages.error(request, "Sale record not found.")
        return redirect('sales')
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('sales')


def delete_sales(request,sales_id):
    sales = Sales.objects.get(id=sales_id)
    sales.delete()
    return redirect ('sales')

def stocks(request):
    stock_info = []
    products = Purchase.objects.all()
    for product in products:
        total_purchased = Purchase.objects.filter(id=product.id).values('quantity').first()['quantity']
        total_sold = Sales.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_returend = Returns.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
        available_stock = total_purchased - total_sold + total_returend
        stock_info.append({'product_name': product.product_name, 'available_stock': available_stock})
    return render(request, 'stocks.html', {'available_stocks': stock_info})




def returns(request):
    try:
        purchased_products = Purchase.objects.all()
        
        if request.method == 'POST':
            product_name = request.POST.get('product_name')
            quantity = request.POST.get('quantity')
            customer = request.POST.get('customer')
            reason = request.POST.get('reason')

            try:
                product = Purchase.objects.filter(product_name=product_name).first()

                if product:
                    Returns.objects.create(
                        product=product, 
                        product_name=product_name, 
                        quantity=int(quantity),  # Ensure valid integer input
                        reason=reason, 
                        customer=customer
                    )
                    messages.success(request, "Return recorded successfully.")
                    return redirect('returns')
                else:
                    messages.error(request, "Product not found in inventory.")

            except ValueError:
                messages.error(request, "Invalid quantity format. Please enter a number.")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {e}")

        return_info = []
        returns = Returns.objects.all()

        for ret in returns:
            product_name = ret.product.product_name 
            total_returned = Returns.objects.filter(product=ret.product).aggregate(Sum('quantity'))['quantity__sum'] or 0  
            return_info.append({
                'product_name': product_name, 
                'total_returned': total_returned, 
                'reason': ret.reason,  
                'quantity': ret.quantity, 
                'customer': ret.customer
            })
        
        return render(request, "returns.html", {
            'return_details': return_info, 
            'purchased_products': purchased_products  
        })
    
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('returns')

    

def new_reports(request):
    leads = []
    try:
        if request.method == "POST":
            from_date = request.POST.get("from_date")
            to_date = request.POST.get("to_date")
            status = request.POST.get("status")

            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

            if from_date and to_date:
                leads = Report.objects.filter(created_date__range=(from_date, to_date))
                if status:
                    leads_count = PersonalInformation.objects.filter(category=status).count()
                    messages.info(request, f"Total leads with status '{status}': {leads_count}")

    except ValueError:
        messages.error(request, "Invalid date format. Please enter dates in YYYY-MM-DD format.")
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")

    return render(request, "new-reports.html", {"leads": leads})
    

def renew_member_list(request):
    all_data = PersonalInformation.objects.filter(status='converted')
    today = date.today()

    for client in all_data:
        service_duration = client.services.duration if client.services else None
        
        if client.assigned_date and service_duration:
            try:
                value, unit = service_duration.split()
                value = int(value)

                if "month" in unit.lower():
                    expiry_date = client.assigned_date + relativedelta(months=value)
                    
                    # Calculate days until expiry
                    days_until_expiry = (expiry_date - today).days
                    
                    # Set notification status
                    if days_until_expiry < 0:
                        notification_status = "Expired"
                    elif days_until_expiry <= 7:
                        notification_status = "Expires this week!"
                    elif days_until_expiry <= 15:
                        notification_status = "Expires in 2 weeks"
                    elif days_until_expiry <= 30:
                        notification_status = "Expires in 1 month"
                    else:
                        notification_status = "Active"
                    
                    client.expiry_date = expiry_date
                    client.notification_status = notification_status
                    client.days_remaining = max(days_until_expiry, 0)
                    
                else:
                    client.expiry_date = None  # ✅ Use None instead of a string
                    client.notification_status = "Error"
                    client.days_remaining = 0
            except ValueError:
                client.expiry_date = None  # ✅ Use None instead of a string
                client.notification_status = "Error"
                client.days_remaining = 0
        else:
            client.expiry_date = None  # ✅ Use None instead of a string
            client.notification_status = "No Plan"
            client.days_remaining = 0

        # ✅ Fix: Use None instead of "Not Renewed Yet"
        client.renewal_date = client.renewal_date if client.renewal_date else None

        # ✅ Ensure expiry_date is stored in the database
        if client.expiry_date:
            client.save()

    return render(request, "renew_member_list.html", {
        "all_data": all_data,
        "today": today,
    })


def renew_member_page(request, member_id):
    try:
        member = get_object_or_404(AddMember, id=member_id)

        if request.method == "POST":
            renewal_date = request.POST.get("renewal_date")

            if not renewal_date:
                messages.error(request, "Please provide a valid renewal date.")
                return redirect("renew_member_page", member_id=member_id)

            member.renewal_date = renewal_date
            member.save()
            messages.success(request, "Membership renewed successfully!")
            return redirect("renew_member_list")

        return render(request, "renew_member_page.html", {"member": member})

    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
        return redirect("renew_member_list")



def renew_member_page(request, member_id):
    try:
        # Fetch the member's details using their ID
        member = get_object_or_404(PersonalInformation, id=member_id)

        if request.method == "POST":
            # Get the renewal date from the form input
            renewal_date = request.POST.get("renewal_date")

            if not renewal_date:
                messages.error(request, "Please provide a valid renewal date.")
                return redirect("renew_member_page", member_id=member_id)

            # Update the renewal date for the member
            member.renewal_date = renewal_date
            member.save()

            # Show success message and redirect
            messages.success(request, "Membership renewed successfully!")
            return redirect("renew_member_list")

        return render(request, "renew_member_page.html", {"member": member})

    except Exception as e:
        messages.error(request, f"An error occurred while renewing the membership: {e}")
        return redirect("renew_member_list")



def save_renewal_date(request, client_id):
    try:
        client = get_object_or_404(PersonalInformation, id=client_id)

        if request.method == 'POST':
            renewal_date_str = request.POST.get('renewal_date')

            if not renewal_date_str:
                messages.error(request, "Please provide a valid renewal date.")
                return redirect('save_renewal_date', client_id=client_id)

            try:
                # Convert the string to a date object
                renewal_date = datetime.strptime(renewal_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
                return redirect('save_renewal_date', client_id=client_id)

            # Save the renewal date to the model
            client.renewal_date = renewal_date
            client.save()

            messages.success(request, "Renewal date updated successfully!")
            return redirect('renew_member_list')

        return render(request, 'renew_member_page.html', {'client': client})

    except Exception as e:
        messages.error(request, f"An error occurred while updating the renewal date: {e}")
        return redirect('renew_member_list')


def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
    except ValueError:
        return None  # Handle invalid date formats

def add_member(request):
    try:
        if request.method == 'POST':
            name = request.POST.get('name')
            gender = request.POST.get('gender')
            mobile_number = request.POST.get('mobile_number')
            aadhar_number = request.POST.get('aadhar_number')
            uploaded_file = request.FILES.get('filename')
            email = request.POST.get('email')
            date_of_birth = parse_date(request.POST.get('date_of_birth', ''))
            location = request.POST.get('location')
            source = request.POST.get('source')
            occupation = request.POST.get('occupation')
            emergency_number = request.POST.get('emergency_number')
            registration_amount = request.POST.get('reg_amount')
            conveniance_fees = float(request.POST.get('conveniance_fees', 0))

            total_amount = request.POST.get('total_amount', '0')
            total_amount = float(total_amount) if total_amount.replace('.', '', 1).isdigit() else 0.0

            enrollment_date = parse_date(request.POST.get('enrollment_date', ''))
            activation_date = parse_date(request.POST.get('activation_date', ''))
            expiry_date = parse_date(request.POST.get('expiry_date', ''))
            payment_date = parse_date(request.POST.get('payment_date', ''))

            current_installment_amount = request.POST.get('current_installment_amount')
            payment_mode = request.POST.get('payment_mode')
            sold_by_id = request.POST.get('sold_by')
            select_membership_plan_id = request.POST.get("select_membership_plan")
            batches_id = request.POST.get("batches")  # Ensure batches_id is correctly fetched

            discount = request.POST.get('discount', '0')
            discount = float(discount) if discount.replace('.', '', 1).isdigit() else 0.0
            discount_type = request.POST.get('discount_type')

            service_id = request.POST.get('service')
            batch_id = request.POST.get('group')
            cost_of_plan_id = request.POST.get('cost_of_plan')
            total_session_id = request.POST.get('total_session')

            service = Service.objects.filter(id=int(service_id)).first() if service_id and service_id.isdigit() else None
            batch = Service.objects.filter(id=int(batch_id)).first() if batch_id and batch_id.isdigit() else None
            cost_of_plan = Service.objects.filter(id=int(cost_of_plan_id)).first() if cost_of_plan_id and cost_of_plan_id.isdigit() else None
            total_session = Service.objects.filter(id=int(total_session_id)).first() if total_session_id and total_session_id.isdigit() else None
            select_membership_plan = Plan.objects.filter(id=int(select_membership_plan_id)).first() if select_membership_plan_id and select_membership_plan_id.isdigit() else None
            
            # Fix: Ensure batches variable is always defined
            batches = Plan.objects.filter(id=int(batches_id)).first() if batches_id and batches_id.isdigit() else None

            sold_by = UserProfile.objects.filter(id=int(sold_by_id)).first() if sold_by_id and sold_by_id.isdigit() else None

            if not name or not gender or not mobile_number or not aadhar_number or not email:
                messages.error(request, "Required fields are missing")
                return redirect('add_member')

            if not sold_by:
                messages.error(request, "Sold By field is required")
                return redirect('add_member')

            member = AddMember(
                name=name,
                gender=gender,
                mobile_number=mobile_number,
                aadhar_number=aadhar_number,
                uploaded_file=uploaded_file,
                email=email,
                date_of_birth=date_of_birth,
                location=location,
                source=source,
                occupation=occupation,
                emergency_number=emergency_number,
                registration_amount=registration_amount,
                service=service,
                batch=batch,
                cost_of_plan=cost_of_plan,
                total_session=total_session,
                conveniance_fees=conveniance_fees,
                total_amount=total_amount,
                enrollment_date=enrollment_date,
                activation_date=activation_date,
                expiry_date=expiry_date,
                current_installment_amount=current_installment_amount,
                payment_mode=payment_mode,
                payment_date=payment_date,
                sold_by=sold_by,
                discount=discount,
                discount_type=discount_type,
                select_membership_plan=select_membership_plan,
                batches=batches,  # Use the fixed batches variable
            )

            member.save()
            messages.success(request, "Member added successfully!")
            return redirect('display_add_members')

        available_services = Service.objects.all()
        group = Service.objects.all()
        sale_by = UserProfile.objects.all()
        membership_plan = Plan.objects.all()

        # Fix: Ensure batches is always defined
        batches = Plan.objects.all()

        return render(request, 'add-members.html', {
            'available_services': available_services,
            'group': group,
            'sale_by': sale_by,
            'membership_plan': membership_plan,
            'batches': batches,  # Pass batches to the template
        })

    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
        return redirect('add_member')



def display_add_members(request):
    add_members = AddMember.objects.all()
    return render(request, "display-all-add-members.html", {"add_members": add_members})



def view_add_members(request, id):
    try:
        obj = get_object_or_404(AddMember, id=id)  # Ensures a 404 if not found

        if request.method == "POST":
            obj.name = request.POST.get('name', obj.name)
            obj.gender = request.POST.get('gender', obj.gender)
            obj.mobile_number = request.POST.get('mobile_number', obj.mobile_number)
            obj.aadhar_number = request.POST.get('aadhar_number', obj.aadhar_number)
            obj.email = request.POST.get('email', obj.email)
            obj.date_of_birth = parse_date(request.POST.get('date_of_birth')) or obj.date_of_birth
            obj.location = request.POST.get('location', obj.location)
            obj.source = request.POST.get('source', obj.source)
            obj.occupation = request.POST.get('occupation', obj.occupation)
            obj.emergency_number = request.POST.get('emergency_number', obj.emergency_number)
            obj.registration_amount = request.POST.get('reg_amount', obj.registration_amount)

            # Convert IDs safely
            service_id = request.POST.get('service')
            batch_id = request.POST.get('group')  # If using 'group' as batch field in form
            cost_of_plan_id = request.POST.get('cost_of_plan')
            total_session_id = request.POST.get('total_session')
            sold_by_id = request.POST.get('sold_by')
            membership_plan_id = request.POST.get('select_membership_plan')

            obj.service = Service.objects.filter(id=int(service_id)).first() if service_id and service_id.isdigit() else obj.service
            obj.batch = Service.objects.filter(id=int(batch_id)).first() if batch_id and batch_id.isdigit() else obj.batch
            obj.cost_of_plan = Service.objects.filter(id=int(cost_of_plan_id)).first() if cost_of_plan_id and cost_of_plan_id.isdigit() else obj.cost_of_plan
            obj.total_session = Service.objects.filter(id=int(total_session_id)).first() if total_session_id and total_session_id.isdigit() else obj.total_session
            obj.sold_by = UserProfile.objects.filter(id=int(sold_by_id)).first() if sold_by_id and sold_by_id.isdigit() else obj.sold_by
            obj.select_membership_plan = Plan.objects.filter(id=int(membership_plan_id)).first() if membership_plan_id and membership_plan_id.isdigit() else obj.select_membership_plan

            # ✅ Fix for Batches (Ensuring it's assigned correctly)
            batches_id = request.POST.get('batches')
            obj.batches = Plan.objects.filter(id=int(batches_id)).first() if batches_id and batches_id.isdigit() else obj.batches

            obj.conveniance_fees = float(request.POST.get('conveniance_fees', obj.conveniance_fees))
            obj.total_amount = float(request.POST.get('total_amount', obj.total_amount))
            obj.enrollment_date = parse_date(request.POST.get('enrollment_date')) or obj.enrollment_date
            obj.activation_date = parse_date(request.POST.get('activation_date')) or obj.activation_date
            obj.expiry_date = parse_date(request.POST.get('expiry_date')) or obj.expiry_date
            obj.current_installment_amount = request.POST.get('current_installment_amount', obj.current_installment_amount)
            obj.payment_mode = request.POST.get('payment_mode', obj.payment_mode)
            obj.payment_date = parse_date(request.POST.get('payment_date')) or obj.payment_date
            obj.discount = float(request.POST.get('discount', obj.discount))
            obj.discount_type = request.POST.get('discount_type', obj.discount_type)

            # File Upload
            uploaded_file = request.FILES.get('filename')
            if uploaded_file:
                obj.uploaded_file = uploaded_file

            obj.save()
            messages.success(request, "Data updated successfully")
            return redirect(f'/display_add_members/{obj.id}/')

        # ✅ Debugging: Print Batches in Console
        print(f"Member: {obj.name}, Batches: {obj.batches}")

        return render(request, "view-each-add-members.html", {"member": obj})

    except ValueError as ve:
        messages.error(request, f"Invalid input: {ve}")
    except Exception as e:
        messages.error(request, f"An error occurred: {e}")

    return redirect('display_add_members')


def delete_added_members(request,id):
    delete_added_members = AddMember.objects.get(id=id)
    delete_added_members.delete()
    return redirect('display_add_members')

def bill(request):
    return render(request,"bill.html")


def view_all_clients(request):
    leads = PersonalInformation.objects.all()
    print(leads)
    add_members = AddMember.objects.all()
    print(add_members)

    return render(request, "all-clients.html", {'leads': leads, 'add_members': add_members})

### REPORTS GENERATION 

def download_report(request):
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    status = request.GET.get('status')

    if not all([from_date, to_date, status]):
        raise BadRequest("Please provide all required parameters")

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Configure report based on category
    if status == 'PersonalInformation':
        ws.title = "Personal Information Report"
        headers = [
            "Name", "Address", "Occupation", "Gender", "Mobile", "Email",
            "Perfect Body", "Body Type", "Goal", "Body You Want",
            "Level of Body Fat", "Problem Areas", "Diets",
            "Water Intake", "Food Division", "Event Coming Up",
            "Height", "Current Weight", "Target Weight", "Level of Fitness"
        ]
        queryset = PersonalInformation.objects.filter(
            created_date__range=[from_date, to_date]
        )
        
    elif status in ['new', 'converted', 'not-converted', 'follow-up', 'pending']:
        ws.title = f"{status.title()} Leads Report"
        headers = [
            "Name", "Mobile", "Email", "Status", "Follow Up Date",
            "Created Date", "Service", "Plan", "Height", "Current Weight",
            "Target Weight", "Goal"
        ]
        queryset = PersonalInformation.objects.filter(
            Q(created_date__range=[from_date, to_date]) &
            Q(status=status)
        )

    elif status == 'sales':
        ws.title = "Sales Report"
        headers = [
            "Product Name", "Customer", "Quantity", "Sale Price",
            "Discount", "Sale Date", "Total Amount"
        ]
        queryset = Sales.objects.filter(sale_date__range=[from_date, to_date])

    elif status == 'purchase':
        ws.title = "Purchase Report"
        headers = [
            "Product Name", "Brand", "Quantity", "Amount",
            "Purchase Date"
        ]
        queryset = Purchase.objects.filter(purchase_date__range=[from_date, to_date])

    elif status == 'returns':
        ws.title = "Returns Report"
        headers = [
            "Product Name", "Customer", "Quantity", "Reason",
            "Return Date"
        ]
        queryset = Returns.objects.filter(return_date__range=[from_date, to_date])

    elif status == 'stocks':
        ws.title = "Stock Report"
        headers = [
            "Product Name", "Quantity", "Purchase Details"
        ]
        queryset = Stock.objects.all()  # Usually stocks don't need date filtering

    elif status == 'payments':
        ws.title = "Payments Report"
        headers = [
            "Member Name", "Amount Paid", "Pending Amount",
            "Payment Mode", "Date Paid", "Created Date"
        ]
        queryset = Payments.objects.filter(date_paid__range=[from_date, to_date])

    elif status == 'members':
        ws.title = "Members Report"
        headers = [
            "Name", "Gender", "Mobile Number", "Email", "Registration Amount", "Service", "Batch", "Cost of Plan", 
            "Current Installment Amount", "Enrollment_date","Membership Plan", "Total Amount"
        ]
        queryset = AddMember.objects.filter(enrollment_date__range = [from_date, to_date]) 

    elif status == 'expenses':
        ws.title = "Expense Report"
        headers = [
            "Expense", "Price", "Quantity", "Date_Spent"
        ]
        queryset = Expense.objects.filter(date_spent__range = [from_date, to_date])


    else:
        raise BadRequest("Invalid report type")

    # Apply headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col)].width = 15

    # Add data based on report type
    for row, item in enumerate(queryset, 2):
        if status == 'PersonalInformation':
            data = [
                item.name, item.address, item.occupation, item.gender,
                item.mobile_number, item.email, item.perfect_body,
                item.body_type, item.goal, item.body_you_want,
                item.level_of_body_fat, item.problem_areas, item.diets,
                item.water_you_drink_daily, item.eat_or_dividie_foods_or_beverages,
                item.event_coming_up, item.height, item.current_weight,
                item.target_weight, item.level_of_fitness
            ]
        elif status in ['new', 'converted', 'not-converted', 'follow-up', 'pending']:
            data = [
                item.name, item.mobile_number, item.email, item.status,
                str(item.follow_up_date), str(item.created_date),
                item.services.name if item.services else '',
                item.plan_leads.plan_name if item.plan_leads else '',
                item.height, item.current_weight, item.target_weight, item.goal
            ]
        elif status == 'sales':
            total_amount = (item.sale_price or 0) * item.quantity
            if item.discount:
                total_amount = total_amount * (1 - item.discount/100)
            data = [
                item.product_name, item.customer, item.quantity,
                item.sale_price, item.discount, str(item.sale_date),
                total_amount
            ]
        elif status == 'purchase':
            data = [
                item.product_name, item.brand, item.quantity,
                item.amount, str(item.purchase_date)
            ]
        elif status == 'returns':
            data = [
                item.product_name, item.customer, item.quantity,
                item.reason, str(item.return_date)
            ]
        elif status == 'stocks':
            data = [
                item.product_name, item.quantity,
                f"Purchase ID: {item.purchase.id}" if item.purchase else 'N/A'
            ]
        elif status == 'payments':
            data = [
                item.name.name if item.name else '', item.amount_paid,
                item.pending_amount, item.payment_mode,
                str(item.date_paid), str(item.created_date)
            ]
        elif status == 'members':
          
           data = [item.name,item.gender,item.mobile_number,item.email,item.registration_amount,
                   item.service.name if item.service else "N/A",item.service.group if item.service else "N/A",
                   item.service.prices if item.service else "N/A", item.current_installment_amount,item.enrollment_date,
                   item.select_membership_plan.plan_name, item.total_amount
                   ]
           
        elif status == 'expenses':
            data = [
                item.expense_name, item.price, item.quantity, item.date_spent
            ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Add summary section
    summary_row = len(queryset) + 3
    ws.cell(row=summary_row, column=1, value="Report Summary").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="Total Records:")
    ws.cell(row=summary_row + 1, column=2, value=len(queryset))
    ws.cell(row=summary_row + 2, column=1, value="Date Range:")
    ws.cell(row=summary_row + 2, column=2, value=f"{from_date} to {to_date}")

    # Add totals for relevant reports
    if status in ['sales', 'payments', 'purchase']:
        total_amount = sum(item.amount_paid if status == 'payments' 
                         else item.amount if status == 'purchase'
                         else (item.sale_price * item.quantity * (1 - (item.discount or 0)/100))
                         for item in queryset)
        ws.cell(row=summary_row + 3, column=1, value="Total Amount:")
        ws.cell(row=summary_row + 3, column=2, value=total_amount)

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename={status}_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

    wb.save(response)
    return response





# import os
# from django.http import HttpResponse
# from django.shortcuts import get_object_or_404
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter
# from reportlab.lib.colors import red, black, blue
# from .models import AddMember

# def generate_invoice(request, member_id):
#     """Generate a styled PDF invoice with a gym logo."""
#     member = get_object_or_404(AddMember, id=member_id)
    
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="invoice_{member.name}.pdf"'
    
#     pdf = canvas.Canvas(response, pagesize=letter)
#     width, height = letter
    
#     # **Gym Logo**
#     logo_path = 'static/image/n1.png'  # Update with actual path
#     if os.path.exists(logo_path):
#         pdf.drawImage(logo_path, 50, height - 110, width=100, height=100)
    
#     # **Header - Gym Name & Contact Details**
#     pdf.setFont("Helvetica-Bold", 22)
#     pdf.setFillColor(blue)
#     pdf.drawCentredString(width / 2, height - 50, "TEFFY FITNESS")
#     pdf.setFillColor(black)
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(50, height - 130, "Phone:  9865638121")
#     pdf.drawString(50, height - 150, "Email: teffyfitness@gmail.com")
#     pdf.drawString(50, height - 170, "GSTIN: -")
    
#     # **Invoice Title**
#     pdf.setFont("Helvetica-Bold", 18)
#     pdf.drawCentredString(width / 2, height - 200, "Payment Receipt")
    
#     # **Member Details**
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawString(50, height - 230, "BILL TO")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(50, height - 250, f"Name: {member.name}")
#     pdf.drawString(50, height - 270, f"BIO ID: {member.id}")
#     pdf.drawString(50, height - 290, f"Phone: {member.mobile_number}")
    
#     # **Invoice Details**
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, height - 250, f"Bill Number: {member.id}")
#     pdf.drawString(300, height - 270, f"Bill Date: {member.enrollment_date}")
#     pdf.drawString(300, height - 290, "Receipt Type: Enrollment")
    
#     # **Purchase Details**
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawString(50, height - 320, "Purchase Details")
#     pdf.setFont("Helvetica", 12)
#     plan_name = member.select_membership_plan.plan_name if member.select_membership_plan else "N/A"
#     pdf.drawString(50, height - 340, f"{plan_name} From {member.enrollment_date} till {member.expiry_date} Rs. {member.total_amount}")
    
#     pdf.drawString(50, height - 360, "SUB-TOTAL")
#     pdf.drawString(300, height - 360, f"Rs. {member.total_amount}")
#     pdf.drawString(50, height - 380, "DISCOUNT")
#     pdf.drawString(300, height - 380, f"Rs. {member.discount if member.discount else 0}")
#     pdf.drawString(50, height - 400, "GRAND TOTAL")
#     pdf.drawString(300, height - 400, f"Rs. {member.total_amount}")
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.setFillColor(red if member.pending_amount() > 0 else black)
#     pdf.drawString(50, height - 420, "DUE AMOUNT")
#     pdf.drawString(300, height - 420, f"Rs. {member.pending_amount()}")
#     pdf.setFillColor(black)
    
#     # **Payment Mode Details**
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawString(50, height - 450, "Payment Mode Details")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(50, height - 470, f"{member.payment_mode}: Rs. {member.current_installment_amount}")
    
#     # **Terms and Conditions**
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawString(50, height - 500, "Terms and Conditions")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(50, height - 520, "No refund policy.")
#     pdf.drawString(50, height - 540, "No disputes or arguments will be entertained without a valid receipt.")
#     pdf.drawString(50, height - 560, "Cheque Dishonor Charges will be applicable in case of cheque bounce.")
#     pdf.drawString(50, height - 580, "Amount of Tax subject to Reverse Charges.")
#     pdf.drawString(50, height - 600, "Please go through the Terms and Conditions pasted at the entrance.")
    
#     # **Finalize and Save the PDF**
#     pdf.showPage()
#     pdf.save()
    
#     return response


def update_member_payment(request, member_id):
    """Handles updating member payments and prevents overpayment."""
    member = get_object_or_404(AddMember, id=member_id)

    if request.method == 'POST':
        amount_paid = float(request.POST.get('amount_paid', 0))
        pending_amount = max(member.total_amount - member.current_installment_amount, 0)

        if amount_paid > pending_amount:
            messages.error(request, f"Payment failed! You cannot pay more than ₹{pending_amount}.")
        elif amount_paid > 0:
            member.current_installment_amount += amount_paid
            member.payment_date = now()

            # Ensure the total paid does not exceed the total amount
            if member.current_installment_amount >= member.total_amount:
                member.current_installment_amount = member.total_amount  # Set it to total amount if overpaid

            member.save()
            pending_amount = max(member.total_amount - member.current_installment_amount, 0)

            messages.success(request, f"Payment of ₹{amount_paid} updated! Pending: ₹{pending_amount}")

    return redirect('view_each_member', member_id=member.id)




def view_each_member(request, member_id):
    """Retrieve a specific member and ensure pending amount is included"""
    member = get_object_or_404(AddMember, id=member_id)
    
    # Fetch the pending amount (same logic as in add-members.html)
    pending_amount = max(member.total_amount - member.current_installment_amount, 0)

    return render(request, 'view-each-add-members.html', {
        'member': member,
        'pending_amount': pending_amount,  # Pass this to the template
    })


# def generate_invoice_pdf(request, client_id):
#     client = get_object_or_404(PersonalInformation, id=client_id)
#     payments = Payments.objects.filter(payments=client) 


#     buffer = BytesIO()
#     pdf = canvas.Canvas(buffer, pagesize=A4)
#     width, height = A4

#     # Business Details
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawString(50, height - 40, "Teffy Sports & Fitness")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(50, height - 55, "2ND FLOOR, JEYABAL COMPLEX, SA ROUND ROAD, Dindigul")
#     pdf.drawString(50, height - 70, "Phone No: 7867800111")
#     pdf.drawString(50, height - 85, "Email: teffysportsacademy@gmail.com")

#     # Business Logo
#     logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'image', 'Mask group (2).png')
#     if os.path.exists(logo_path):
#         pdf.drawImage(logo_path, width - 100, height - 100, width=80, height=80)

#     # Tax Invoice Title
#     pdf.setFont("Helvetica-Bold", 16)
#     pdf.drawCentredString(width / 2, height - 120, "Tax Invoice")

#     # Client Information
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(50, height - 150, "Bill To")
#     pdf.setFont("Helvetica", 11)
#     pdf.drawString(50, height - 170, f"{client.name}")
#     pdf.drawString(50, height - 190, f"Contact No: {client.mobile_number}")

#     # Invoice Details (Right Side)
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(width - 200, height - 150, "Invoice Details")
#     pdf.setFont("Helvetica", 11)
#     pdf.drawString(width - 200, height - 170, f"Invoice No: {client.id}")
#     # pdf.drawString(width - 200, height - 190, f"Date: {client.created_at.strftime('%d-%m-%Y')}")

#     # Service Details
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(50, height - 220, "Description")
#     pdf.setFont("Helvetica", 11)
#     pdf.drawString(50, height - 240, f"{client.plan_leads.plan_name if client.plan_leads else 'N/A'}")

#     # Payment Summary Table
#     total_paid = sum(payment.amount_paid for payment in payments)
#     pending_amount = payments.last().pending_amount if payments.exists() else 0

#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(50, height - 280, "Total")
#     pdf.drawString(50, height - 300, "Received")
#     pdf.drawString(50, height - 320, "Balance")

#     pdf.setFont("Helvetica", 11)
#     pdf.drawString(150, height - 280, f"₹ {total_paid:,.2f}")
#     pdf.drawString(150, height - 300, f"₹ {total_paid:,.2f}")
#     pdf.drawString(150, height - 320, f"₹ {pending_amount:,.2f}")

#     # Authorized Signatory
#     pdf.line(50, height - 360, 200, height - 360)
#     pdf.drawString(50, height - 375, "For: Teffy Sports & Fitness")
#     pdf.drawString(50, height - 390, "Authorized Signatory")

#     pdf.showPage()
#     pdf.save()

#     buffer.seek(0)
#     response = HttpResponse(buffer, content_type="application/pdf")
#     response["Content-Disposition"] = f'attachment; filename="invoice_{client.name}.pdf"'
#     return response


# from django.shortcuts import render, get_object_or_404
# from django.http import HttpResponse
# from weasyprint import HTML
# from models import *
# import datetime

# def generate_invoice_pdf(request, client_id):
#     client = get_object_or_404(PersonalInformation, id=client_id)
#     payments = Payments.objects.filter(client=client)
    
#     total_paid = sum(payment.amount_paid for payment in payments)
#     pending_amount = payments.last().pending_amount if payments.exists() else 0

#     context = {
#         "client": client,
#         "bill_number": f"INV-{client.id:04d}",  # Generates invoice number like INV-0001
#         "bill_date": datetime.date.today().strftime("%d/%m/%Y"),
#         "payments": payments,
#         "total_paid": total_paid,
#         "pending_amount": pending_amount,
#     }

#     # Render HTML
#     html_string = render(request, "invoice_template.html", context).content.decode("utf-8")
#     pdf_file = HTML(string=html_string).write_pdf()

#     # Serve the PDF
#     response = HttpResponse(pdf_file, content_type="application/pdf")
#     response["Content-Disposition"] = f'attachment; filename="invoice_{client.name}.pdf"'
#     return response


# from django.shortcuts import render, get_object_or_404
# from django.http import HttpResponse
# from weasyprint import HTML
# from django.template.loader import render_to_string
# from .models import AddMember

# def generate_invoice_pdf(request, member_id):
#     member = get_object_or_404(AddMember, id=member_id)

#     # Render invoice as an HTML string
#     html_content = render_to_string("invoice.html", {"member": member})

#     # Convert HTML to PDF
#     pdf = HTML(string=html_content).write_pdf()

#     # Create HTTP response to serve the file
#     response = HttpResponse(pdf, content_type="application/pdf")
#     response["Content-Disposition"] = f'attachment; filename="Invoice_{member.name}.pdf"'

#     return response

### INVOICE GENERATION ###

# ✅ Generate Invoice (HTML) for Both Leads and Members
def generate_invoice(request, client_id, client_type):
    if client_type == "lead":
        client = get_object_or_404(PersonalInformation, id=client_id)
        service = client.services
        plan = client.plan_leads
        total_paid = Payments.objects.filter(payments=client).aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0
    else:
        client = get_object_or_404(AddMember, id=client_id)
        service = client.service
        plan = client.select_membership_plan
        total_paid = client.current_installment_amount if client.current_installment_amount else 0

    service_price = service.prices if service else 0
    plan_price = plan.price if plan else 0
    total_amount = service_price + plan_price
    total_due = max(total_amount - total_paid, 0)

    context = {
        "client": client,
        "bill_number": client.id,
        "bill_date": now().date(),
        "service": service,
        "plan": plan,
        "total_amount": total_amount,
        "paid_amount": total_paid,
        "due_amount": total_due,
        "receipt_type": "Converted Lead" if client_type == "lead" else "Direct Member",
        "client_type": client_type,
        "logo_url": request.build_absolute_uri(static("image/n1.png")),
    }

    template = get_template("invoice.html")
    html = template.render(context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice_{client.id}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response


# ✅ Generate Invoice PDF for Members
def generate_invoice_pdf_member(request, member_id):
    """ Generate invoice PDF for a member """

    # Fetch member details
    member = get_object_or_404(AddMember, id=member_id)

    # Get service and membership details
    service = member.service
    batch = member.batches.batch if member.batches else "N/A"
    expiry_date = member.expiry_date if member.expiry_date else "N/A"
    discount = member.discount or 0

    # ✅ Get total amount correctly
    service_price = service.prices if service else 0
    total_amount = max(service_price - discount, 0)  # Apply discount

    # ✅ Get total paid amount directly from `current_installment_amount`
    total_paid = member.current_installment_amount or 0

    # ✅ Calculate correct pending amount
    total_due = max(total_amount - total_paid, 0)

    # Prepare context for the invoice template
    context = {
        "client": member,
        "bill_number": member.id,
        "bill_date": now().date(),
        "service": service,
        "total_amount": total_amount,
        "paid_amount": total_paid,
        "due_amount": total_due,
        "receipt_type": "Direct Member",
        "client_type": "member",
        "logo_url": request.build_absolute_uri(static("image/n1.png")),
        "expiry_date": expiry_date,
        "batch": batch,
    }

    # Load invoice template
    template = get_template("invoice.html")
    html = template.render(context)

    # Generate PDF response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=\"invoice_{member.id}.pdf\"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response


# ✅ Generate Invoice PDF for Leads
def generate_invoice_pdf_lead(request, lead_id):
    lead = get_object_or_404(PersonalInformation, id=lead_id)
    service = lead.services

    # ✅ Fetch expiry date
    expiry_date = lead.expiry_date if lead.expiry_date else "Not Available"

    # ✅ Get Total Paid Amount
    payments = Payments.objects.filter(payments=lead)
    total_paid = payments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0

    # ✅ Calculate Total Amount (After Discount)
    service_price = service.prices if service else 0
    discount = lead.discount if lead.discount else 0
    total_amount = max(service_price - discount, 0)

    # ✅ Calculate Pending Amount
    total_due = max(total_amount - total_paid, 0)

    context = {
        "client": lead,
        "bill_number": lead.id,
        "bill_date": now().date(),
        "service": service,
        "total_amount": total_amount,
        "paid_amount": total_paid,
        "due_amount": total_due,
        "receipt_type": "Converted Lead",
        "client_type": "lead",
        "logo_url": request.build_absolute_uri(static("image/n1.png")),
        "expiry_date": expiry_date,  # ✅ Pass expiry date to the template
    }

    template = get_template("invoice.html")
    html = template.render(context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice_{lead.id}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response
